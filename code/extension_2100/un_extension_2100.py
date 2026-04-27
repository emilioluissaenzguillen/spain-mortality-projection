"""Extend the validated INE-style mortality replication from 2073 to 2100.

The extension keeps the current replication unchanged through 2073 and then
uses UN life-expectancy projections for Spain to define the annual target
path from 2074 to 2100.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import importlib.util
import math
import re
import sys

import pandas as pd
from openpyxl import Workbook, load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]
REPLICATION_SCRIPT = PROJECT_ROOT / "code" / "replication" / "ine_methodology_replication.py"
UN_MALE_E0_FILE = PROJECT_ROOT / "input" / "demographic_inputs" / "un_extension" / "wpp2019_life_expectancy_at_birth_male.xlsx"
UN_FEMALE_E0_FILE = PROJECT_ROOT / "input" / "demographic_inputs" / "un_extension" / "wpp2019_life_expectancy_at_birth_female.xlsx"

FINAL_DIR = PROJECT_ROOT / "output" / "mortality_projection" / "final" / "extension_2100"
VALIDATION_FILE = FINAL_DIR / "un_extension_validation_2074_2100.xlsx"

BASE_VARIANT = "baseline"
BASE_END_YEAR = 2073
EXTENSION_START_YEAR = 2074
EXTENSION_END_YEAR = 2100


@dataclass(frozen=True)
class SexConfig:
    sex: str
    model_family: str
    un_file: Path


SEX_CONFIGS = [
    SexConfig(sex="male", model_family="East", un_file=UN_MALE_E0_FILE),
    SexConfig(sex="female", model_family="West", un_file=UN_FEMALE_E0_FILE),
]


def require_file(path: Path) -> None:
    if not path.exists():
        raise FileNotFoundError(f"Missing required file: {path}")


def ensure_dirs() -> None:
    FINAL_DIR.mkdir(parents=True, exist_ok=True)


def load_replication_module():
    module_name = "ine_methodology_replication"
    spec = importlib.util.spec_from_file_location(module_name, REPLICATION_SCRIPT)
    if spec is None or spec.loader is None:
        raise ImportError(f"Could not load replication module from: {REPLICATION_SCRIPT}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[module_name] = module
    spec.loader.exec_module(module)
    return module


def parse_un_spain_period_e0(path: Path) -> dict[int, float]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["MEDIUM VARIANT"]

    header = [ws.cell(17, col).value for col in range(1, ws.max_column + 1)]
    row_values = None
    for row in ws.iter_rows(min_row=18, values_only=True):
        if row[2] == "Spain":
            row_values = list(row)
            break

    if row_values is None:
        raise ValueError(f"Could not find Spain in UN workbook: {path}")

    period_end_e0: dict[int, float] = {}
    for idx in range(7, len(header)):
        label = header[idx]
        value = row_values[idx]
        if label is None or value is None:
            continue
        match = re.fullmatch(r"(\d{4})-(\d{4})", str(label))
        if match is None:
            continue
        period_end_year = int(match.group(2))
        period_end_e0[period_end_year] = float(value)

    return period_end_e0


def build_annual_un_target_path(start_e0_2073: float, period_end_e0: dict[int, float]) -> pd.DataFrame:
    knots = {BASE_END_YEAR: start_e0_2073}
    knots.update({year: value for year, value in period_end_e0.items() if 2075 <= year <= EXTENSION_END_YEAR})
    sorted_years = sorted(knots)

    rows = [{"Year": BASE_END_YEAR, "target_e0": start_e0_2073, "source": "replication_2073"}]
    for left_year, right_year in zip(sorted_years[:-1], sorted_years[1:]):
        left_value = knots[left_year]
        right_value = knots[right_year]
        for year in range(left_year + 1, right_year + 1):
            fraction = (year - left_year) / (right_year - left_year)
            target = left_value + fraction * (right_value - left_value)
            source = f"linear_{left_year}_{right_year}"
            if year == right_year and year in period_end_e0:
                source = f"un_period_end_{year}"
            rows.append({"Year": year, "target_e0": target, "source": source})

    return pd.DataFrame(rows)


def build_extension_coefficients(target_path: pd.DataFrame) -> dict[int, float]:
    target_by_year = {
        int(row.Year): float(row.target_e0)
        for row in target_path.itertuples(index=False)
    }
    horizon_e0 = target_by_year[EXTENSION_END_YEAR]

    coefficients: dict[int, float] = {}
    for year in range(EXTENSION_START_YEAR, EXTENSION_END_YEAR):
        previous_e0 = target_by_year[year - 1]
        current_e0 = target_by_year[year]
        coefficients[year] = (current_e0 - previous_e0) / (horizon_e0 - previous_e0)

    return coefficients


def interpolate_horizon_profiles(replication, sex: str, family: str, target_e0_2100: float) -> tuple[dict[int, float], dict[int, float]]:
    lower_e0 = math.floor(target_e0_2100)
    upper_e0 = lower_e0 + 1
    coefficient = target_e0_2100 - lower_e0

    qx_profiles, ax_profiles = replication.parse_model_profiles(sex, family, lower_e0, upper_e0)
    qx_horizon = replication.interpolate_profiles(qx_profiles[lower_e0], qx_profiles[upper_e0], coefficient)
    qx_horizon[100] = 1.0
    ax_horizon = replication.interpolate_profiles(ax_profiles[lower_e0], ax_profiles[upper_e0], coefficient)
    return qx_horizon, ax_horizon


def project_extension_profiles(
    start_profile: dict[int, float],
    horizon_profile: dict[int, float],
    coefficients: dict[int, float],
    *,
    force_terminal_qx: bool,
) -> pd.DataFrame:
    ages = sorted(start_profile)
    previous = start_profile.copy()
    projected: dict[int, dict[int, float]] = {}

    for year in range(EXTENSION_START_YEAR, EXTENSION_END_YEAR):
        coefficient = coefficients[year]
        current = {
            age: coefficient * horizon_profile[age] + (1 - coefficient) * previous[age]
            for age in ages
        }
        if force_terminal_qx:
            current[100] = 1.0
        projected[year] = current
        previous = current

    projected[EXTENSION_END_YEAR] = {
        age: (1.0 if force_terminal_qx and age == 100 else value)
        for age, value in horizon_profile.items()
    }

    df = pd.DataFrame({"Age": ages})
    for year in range(EXTENSION_START_YEAR, EXTENSION_END_YEAR + 1):
        df[f"Year_{year}"] = [projected[year][age] for age in ages]
    return df


def build_extension_life_projection(replication, qx_extension: pd.DataFrame, ax_extension: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, float | int]] = []
    for year in range(EXTENSION_START_YEAR, EXTENSION_END_YEAR + 1):
        qx_by_age = {
            int(row.Age): float(getattr(row, f"Year_{year}"))
            for row in qx_extension.itertuples(index=False)
        }
        ax_by_age = {
            int(row.Age): float(getattr(row, f"Year_{year}"))
            for row in ax_extension.itertuples(index=False)
        }
        rows.append(
            {
                "Year": year,
                "e0_life_table": replication.life_expectancy_at_birth(qx_by_age, ax_by_age),
            }
        )
    return pd.DataFrame(rows)


def merge_yearly_tables(base_df: pd.DataFrame, extension_df: pd.DataFrame) -> pd.DataFrame:
    merged = base_df.copy()
    for year in range(EXTENSION_START_YEAR, EXTENSION_END_YEAR + 1):
        merged[f"Year_{year}"] = extension_df[f"Year_{year}"]
    return merged


def merge_life_projection(base_df: pd.DataFrame, extension_df: pd.DataFrame) -> pd.DataFrame:
    base_rows = base_df[base_df["Year"] <= BASE_END_YEAR].copy()
    extension_rows = extension_df[extension_df["Year"] >= EXTENSION_START_YEAR].copy()
    return pd.concat([base_rows, extension_rows], ignore_index=True)


def write_sheet(ws, headers: list[str], rows: list[list] | list[tuple] | list[dict]) -> None:
    ws.append(headers)
    for row in rows:
        if isinstance(row, dict):
            ws.append([row.get(header) for header in headers])
        else:
            ws.append(list(row))


def save_dataframe(df: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(path, index=False)


def run_extension(replication) -> dict[str, dict[str, pd.DataFrame]]:
    observed_qx, observed_ax, observed_e0, benchmark_qx, benchmark_e0 = replication.load_all_inputs()
    replication_results = replication.run_replication_pipeline(observed_qx, observed_ax, observed_e0)
    base_results = replication_results[BASE_VARIANT]

    extension_results: dict[str, dict[str, pd.DataFrame]] = {}

    for config in SEX_CONFIGS:
        base_result = base_results[config.sex]
        start_e0_2073 = float(
            base_result.life_projection.loc[base_result.life_projection["Year"] == BASE_END_YEAR, "e0_life_table"].iloc[0]
        )
        un_period_end_e0 = parse_un_spain_period_e0(config.un_file)
        target_path = build_annual_un_target_path(start_e0_2073, un_period_end_e0)
        coefficients = build_extension_coefficients(target_path)

        qx_horizon, ax_horizon = interpolate_horizon_profiles(
            replication,
            config.sex,
            config.model_family,
            float(target_path.loc[target_path["Year"] == EXTENSION_END_YEAR, "target_e0"].iloc[0]),
        )

        start_qx = {
            int(row.Age): float(getattr(row, f"Year_{BASE_END_YEAR}"))
            for row in base_result.qx_projected.itertuples(index=False)
        }
        start_ax = {
            int(row.Age): float(getattr(row, f"Year_{BASE_END_YEAR}"))
            for row in base_result.ax_projected.itertuples(index=False)
        }

        qx_extension = project_extension_profiles(start_qx, qx_horizon, coefficients, force_terminal_qx=True)
        ax_extension = project_extension_profiles(start_ax, ax_horizon, coefficients, force_terminal_qx=False)
        life_extension = build_extension_life_projection(replication, qx_extension, ax_extension)
        life_extension = life_extension.merge(target_path[["Year", "target_e0"]], on="Year", how="left")
        life_extension["abs_error_vs_target"] = (life_extension["e0_life_table"] - life_extension["target_e0"]).abs()

        full_qx = merge_yearly_tables(base_result.qx_projected, qx_extension)
        full_ax = merge_yearly_tables(base_result.ax_projected, ax_extension)
        full_life = merge_life_projection(base_result.life_projection, life_extension[["Year", "e0_life_table"]])

        extension_results[config.sex] = {
            "target_path": target_path,
            "qx_horizon": pd.DataFrame({"Age": sorted(qx_horizon), "qx_proj2100": [qx_horizon[age] for age in sorted(qx_horizon)]}),
            "ax_horizon": pd.DataFrame({"Age": sorted(ax_horizon), "ax_proj2100": [ax_horizon[age] for age in sorted(ax_horizon)]}),
            "qx_extension": qx_extension,
            "ax_extension": ax_extension,
            "life_extension": life_extension,
            "qx_full_2024_2100": full_qx,
            "ax_full_2024_2100": full_ax,
            "life_full_2024_2100": full_life,
        }

    return extension_results


def write_extension_outputs(results: dict[str, dict[str, pd.DataFrame]]) -> None:
    for sex, result in results.items():
        save_dataframe(result["target_path"], FINAL_DIR / f"un_e0_target_path_{sex}_2073_2100.xlsx")
        save_dataframe(result["qx_horizon"], FINAL_DIR / f"un_qx_horizon_2100_{sex}.xlsx")
        save_dataframe(result["ax_horizon"], FINAL_DIR / f"un_ax_horizon_2100_{sex}.xlsx")
        save_dataframe(result["qx_extension"], FINAL_DIR / f"un_qx_extension_2074_2100_{sex}.xlsx")
        save_dataframe(result["ax_extension"], FINAL_DIR / f"un_ax_extension_2074_2100_{sex}.xlsx")
        save_dataframe(result["life_extension"], FINAL_DIR / f"un_e0_extension_2074_2100_{sex}.xlsx")
        save_dataframe(result["qx_full_2024_2100"], FINAL_DIR / f"un_qx_2024_2100_{sex}.xlsx")
        save_dataframe(result["ax_full_2024_2100"], FINAL_DIR / f"un_ax_2024_2100_{sex}.xlsx")
        save_dataframe(result["life_full_2024_2100"], FINAL_DIR / f"un_e0_life_table_2024_2100_{sex}.xlsx")


def write_validation_workbook(results: dict[str, dict[str, pd.DataFrame]]) -> None:
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "summary"
    write_sheet(
        summary_ws,
        ["sex", "metric", "compared_cells", "mean_abs_error", "max_abs_error"],
        [],
    )

    for sex, result in results.items():
        life_extension = result["life_extension"]
        abs_errors = life_extension["abs_error_vs_target"].tolist()
        summary_ws.append(
            [
                sex,
                "e0_life_table_vs_annualized_un_target_2074_2100",
                len(abs_errors),
                sum(abs_errors) / len(abs_errors),
                max(abs_errors),
            ]
        )

        detail_ws = wb.create_sheet(f"{sex}_e0")
        write_sheet(
            detail_ws,
            ["Year", "target_e0", "e0_life_table", "abs_error_vs_target"],
            life_extension[["Year", "target_e0", "e0_life_table", "abs_error_vs_target"]].to_dict("records"),
        )

        qx_ws = wb.create_sheet(f"{sex}_qx_2100")
        write_sheet(
            qx_ws,
            ["Age", f"Year_{EXTENSION_END_YEAR}"],
            result["qx_extension"][["Age", f"Year_{EXTENSION_END_YEAR}"]].to_dict("records"),
        )

    notes_ws = wb.create_sheet("method_notes")
    write_sheet(
        notes_ws,
        ["topic", "note"],
        [
            [
                "baseline",
                "The extension keeps the INE-style replication unchanged through 2073 and starts the UN-driven layer in 2074.",
            ],
            [
                "un annualization",
                "UN quinquennial life expectancy values are treated as knots at the period end years 2075, 2080, ..., 2100 and annualized by linear interpolation.",
            ],
            [
                "projection mechanics",
                "Age-specific qx and ax profiles are extended from the replicated 2073 endpoint to a 2100 horizon profile using the same interpolation mechanics as the INE workflow.",
            ],
        ],
    )

    wb.save(VALIDATION_FILE)


def main() -> None:
    for path in [REPLICATION_SCRIPT, UN_MALE_E0_FILE, UN_FEMALE_E0_FILE]:
        require_file(path)
    ensure_dirs()

    replication = load_replication_module()
    results = run_extension(replication)
    write_extension_outputs(results)
    write_validation_workbook(results)

    print(f"Wrote extension outputs to: {FINAL_DIR}")
    print(f"Wrote extension validation workbook: {VALIDATION_FILE}")


if __name__ == "__main__":
    main()
