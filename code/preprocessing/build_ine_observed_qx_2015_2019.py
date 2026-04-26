from __future__ import annotations

from pathlib import Path
import re

import pandas as pd
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]
SOURCE_FILE = (
    PROJECT_ROOT
    / "input"
    / "demographic_inputs"
    / "fresh_downloads"
    / "ine_27153_mortality_tables_functions.xlsx"
)
OUTPUT_DIR = (
    PROJECT_ROOT
    / "output"
    / "mortality_projection"
    / "intermediate"
    / "ine_replication"
    / "observed_inputs"
)


def parse_observed_rows() -> list[dict[str, float | int | str]]:
    wb = load_workbook(SOURCE_FILE, read_only=True, data_only=True)
    ws = wb["tabla-27153"]

    rows: list[dict[str, float | int | str]] = []
    sex: str | None = None
    age: int | None = None

    for row in ws.iter_rows(values_only=True):
        label = row[0]
        if label is None:
            continue

        text = str(label).strip()
        if text == "Hombres":
            sex = "male"
            age = None
            continue
        if text == "Mujeres":
            sex = "female"
            age = None
            continue
        if text == "Ambos sexos":
            sex = "both"
            age = None
            continue

        if row[1] is None and text and any(ch.isdigit() for ch in text):
            match = re.search(r"\d+", text)
            if match:
                age = int(match.group())
            continue

        if sex not in {"male", "female"} or age is None or not re.fullmatch(r"\d{4}", text):
            continue

        year = int(text)
        if not 2015 <= year <= 2019:
            continue

        qx_per_thousand = row[3]
        ax = row[2]
        if qx_per_thousand is None or ax is None:
            continue

        rows.append(
            {
                "sex": sex,
                "age": age,
                "year": year,
                "qx": float(qx_per_thousand) / 1000.0,
                "ax": float(ax),
            }
        )

    return rows


def build_wide_series(df: pd.DataFrame, value_column: str) -> pd.DataFrame:
    wide = (
        df.pivot(index="age", columns="year", values=value_column)
        .rename(columns=lambda year: f"{value_column}_{year}")
        .reset_index()
        .rename(columns={"age": "Age"})
        .sort_values("Age")
    )
    return wide


def build_twice_smoothed_qx(qx_wide: pd.DataFrame) -> pd.DataFrame:
    smoothed = qx_wide.copy()
    smoothed["qxprime_2019"] = (smoothed["qx_2017"] + smoothed["qx_2018"] + 3 * smoothed["qx_2019"]) / 5
    smoothed["qxprime_2017"] = (
        smoothed["qx_2015"] + smoothed["qx_2016"] + smoothed["qx_2017"] + smoothed["qx_2018"] + smoothed["qx_2019"]
    ) / 5
    smoothed["qxprime_2018"] = (smoothed["qx_2016"] + smoothed["qx_2017"] + smoothed["qx_2018"] + 2 * smoothed["qx_2019"]) / 5
    smoothed["qx_2019_twice_smoothed"] = (
        smoothed["qxprime_2017"] + smoothed["qxprime_2018"] + 3 * smoothed["qxprime_2019"]
    ) / 5
    return smoothed[["Age", "qx_2019_twice_smoothed"]]


def write_outputs(rows: list[dict[str, float | int | str]]) -> list[Path]:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(rows)
    written_files: list[Path] = []

    for sex in ["male", "female"]:
        subset = df[df["sex"] == sex].copy()
        qx_wide = build_wide_series(subset, "qx")
        ax_wide = build_wide_series(subset, "ax")
        smoothed_qx = build_twice_smoothed_qx(qx_wide)

        qx_path = OUTPUT_DIR / f"ine_observed_qx_2015_2019_{sex}.csv"
        ax_path = OUTPUT_DIR / f"ine_observed_ax_2015_2019_{sex}.csv"
        smoothed_qx_path = OUTPUT_DIR / f"ine_qx_2019_twice_smoothed_{sex}.xlsx"

        qx_wide.to_csv(qx_path, index=False)
        ax_wide.to_csv(ax_path, index=False)
        smoothed_qx.to_excel(smoothed_qx_path, index=False)
        written_files.extend([qx_path, ax_path, smoothed_qx_path])

    return written_files


def main() -> None:
    if not SOURCE_FILE.exists():
        raise FileNotFoundError(f"Missing required source file: {SOURCE_FILE}")

    rows = parse_observed_rows()
    if not rows:
        raise ValueError(f"No observed qx/ax rows were extracted from: {SOURCE_FILE}")

    written_files = write_outputs(rows)
    for path in written_files:
        print(path)


if __name__ == "__main__":
    main()
