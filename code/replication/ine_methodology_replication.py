from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import math
import re

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]
OBSERVED_FILE = (
    PROJECT_ROOT
    / "input"
    / "demographic_inputs"
    / "fresh_downloads"
    / "ine_27153_mortality_tables_functions.xlsx"
)
MODEL_TABLE_FILE = PROJECT_ROOT / "input" / "demographic_inputs" / "ine_replication" / "MLT_UN2011_130_1y_complete.xlsx"
BENCHMARK_FILE = PROJECT_ROOT / "input" / "published_benchmarks" / "fresh_downloads" / "36774.xlsx"
BENCHMARK_E0_FILE = PROJECT_ROOT / "input" / "published_benchmarks" / "fresh_downloads" / "36775.xlsx"
POPULATION_EXTRACT_CSV = (
    PROJECT_ROOT / "input" / "demographic_inputs" / "fresh_downloads" / "ine_56934_male_july_2023_population_by_age.csv"
)
DEATHS_EXTRACT_CSV = (
    PROJECT_ROOT / "input" / "demographic_inputs" / "fresh_downloads" / "ine_31912_male_2023_total_deaths_by_age.csv"
)

INTERMEDIATE_DIR = PROJECT_ROOT / "output" / "mortality_projection" / "intermediate" / "ine_replication_2024_2073"
FINAL_DIR = PROJECT_ROOT / "output" / "mortality_projection" / "final" / "ine_replication_2024_2073"
VALIDATION_FILE = FINAL_DIR / "ine_replication_validation_2024_2073.xlsx"

START_YEAR = 2024
END_YEAR = 2073
LAST_OBSERVED_YEAR = 2023
OBSERVED_E0_START_YEAR = 1991
SMOOTHING_YEARS = [2015, 2016, 2017, 2018, 2019]
EXCLUDED_REGRESSION_YEARS = {2020, 2021}
CURRENT_START_PROFILE_YEARS = [2017, 2018, 2019, 2022, 2023]
VARIANT_BASELINE = "baseline"
VARIANT_MALE_HIGH_AGE_ADJUSTED = "male_high_age_adjusted"
VARIANT_SHEET_PREFIX = {
    VARIANT_BASELINE: "base",
    VARIANT_MALE_HIGH_AGE_ADJUSTED: "adj",
}


@dataclass(frozen=True)
class SexConfig:
    sex: str
    model_family: str
    benchmark_label: str
    horizon_target_e0: float
    gap_distribution_years: int


CONFIGS = [
    SexConfig(
        sex="male",
        model_family="East",
        benchmark_label="Males",
        horizon_target_e0=86.0,
        gap_distribution_years=40,
    ),
    SexConfig(
        sex="female",
        model_family="West",
        benchmark_label="Females",
        horizon_target_e0=90.0,
        gap_distribution_years=20,
    ),
]


@dataclass
class SummaryRow:
    sex: str
    metric: str
    compared_cells: int
    mean_abs_error: float
    max_abs_error: float


def require_file(path: Path) -> None:
    if not path.exists():
        raise FileNotFoundError(f"Missing required file: {path}")


def ensure_dirs() -> None:
    INTERMEDIATE_DIR.mkdir(parents=True, exist_ok=True)
    FINAL_DIR.mkdir(parents=True, exist_ok=True)


def parse_age_label(text: str) -> int | None:
    if re.fullmatch(r"\d{4}", text.strip()):
        return None
    match = re.search(r"\d+", text)
    if match is None:
        return None
    return int(match.group())


def parse_observed_ine_table() -> tuple[
    dict[tuple[str, int, int], float],
    dict[tuple[str, int, int], float],
    dict[tuple[str, int], float],
]:
    wb = load_workbook(OBSERVED_FILE, read_only=True, data_only=True)
    ws = wb["tabla-27153"]

    observed_qx: dict[tuple[str, int, int], float] = {}
    observed_ax: dict[tuple[str, int, int], float] = {}
    observed_e0: dict[tuple[str, int], float] = {}

    current_sex: str | None = None
    current_age: int | None = None

    for row in ws.iter_rows(values_only=True):
        label = row[0]
        if label is None:
            continue

        text = str(label).strip()
        if text == "Hombres":
            current_sex = "male"
            current_age = None
            continue
        if text == "Mujeres":
            current_sex = "female"
            current_age = None
            continue
        if text == "Ambos sexos":
            current_sex = "both"
            current_age = None
            continue

        if row[1] is None and text:
            age = parse_age_label(text)
            if age is not None:
                current_age = age
            continue

        if (
            current_sex not in {"male", "female"}
            or current_age is None
            or not re.fullmatch(r"\d{4}", text)
        ):
            continue

        year = int(text)
        qx_per_thousand = row[3]
        ax = row[2]
        ex = row[8] if len(row) > 8 else None

        if qx_per_thousand is not None:
            observed_qx[(current_sex, current_age, year)] = float(qx_per_thousand) / 1000.0
        if ax is not None:
            observed_ax[(current_sex, current_age, year)] = float(ax)
        if current_age == 0 and ex is not None:
            observed_e0[(current_sex, year)] = float(ex)

    return observed_qx, observed_ax, observed_e0


def build_observed_wide(
    observed: dict[tuple[str, int, int], float],
    sex: str,
    years: list[int],
) -> pd.DataFrame:
    rows: list[dict[str, float | int]] = []
    ages = sorted({age for current_sex, age, year in observed if current_sex == sex and year in years})
    for age in ages:
        row: dict[str, float | int] = {"Age": age}
        for year in years:
            row[f"value_{year}"] = observed[(sex, age, year)]
        rows.append(row)
    return pd.DataFrame(rows)


def two_pass_moving_average(df: pd.DataFrame) -> pd.DataFrame:
    smoothed = df.copy()
    smoothed["prime_2019"] = (smoothed["value_2017"] + smoothed["value_2018"] + 3 * smoothed["value_2019"]) / 5
    smoothed["prime_2017"] = (
        smoothed["value_2015"]
        + smoothed["value_2016"]
        + smoothed["value_2017"]
        + smoothed["value_2018"]
        + smoothed["value_2019"]
    ) / 5
    smoothed["prime_2018"] = (
        smoothed["value_2016"]
        + smoothed["value_2017"]
        + smoothed["value_2018"]
        + 2 * smoothed["value_2019"]
    ) / 5
    result = smoothed[["Age"]].copy()
    result["smoothed_2019"] = (smoothed["prime_2017"] + smoothed["prime_2018"] + 3 * smoothed["prime_2019"]) / 5
    return result


def build_current_start_profile(
    observed: dict[tuple[str, int, int], float],
    sex: str,
) -> pd.DataFrame:
    rows: list[dict[str, float | int]] = []
    for age in range(0, 101):
        v2017 = observed[(sex, age, 2017)]
        v2018 = observed[(sex, age, 2018)]
        v2019 = observed[(sex, age, 2019)]
        v2022 = observed[(sex, age, 2022)]
        v2023 = observed[(sex, age, 2023)]

        prime_2023 = (v2019 + v2022 + 3 * v2023) / 5
        prime_2022 = (v2018 + v2019 + v2022 + 2 * v2023) / 5
        prime_2019 = (v2017 + v2018 + v2019 + v2022 + v2023) / 5

        rows.append(
            {
                "Age": age,
                "smoothed_start": (prime_2019 + prime_2022 + 3 * prime_2023) / 5,
            }
        )
    return pd.DataFrame(rows)


def parse_number(text: str) -> float:
    return float(text.replace(".", "").replace(",", "."))


def load_population_male_july_2023() -> dict[int, float]:
    values: dict[int, float] = {}
    with POPULATION_EXTRACT_CSV.open(encoding="utf-8-sig", newline="") as handle:
        import csv

        reader = csv.DictReader(handle)
        for row in reader:
            values[int(row["Age"])] = float(row["Population"])
    return values


def load_deaths_male_2023() -> dict[int, float]:
    values: dict[int, float] = {}
    with DEATHS_EXTRACT_CSV.open(encoding="utf-8-sig", newline="") as handle:
        import csv

        reader = csv.DictReader(handle)
        for row in reader:
            values[int(row["Age"])] = float(row["Deaths"])
    return values


def estimated_qx_2023_male(
    population: dict[int, float],
    deaths: dict[int, float],
    observed_qx: dict[tuple[str, int, int], float],
) -> dict[int, float]:
    qx: dict[int, float] = {}
    qx[0] = observed_qx[("male", 0, 2019)]
    for age in range(1, 100):
        mx = deaths[age] / population[age]
        qx[age] = (2 * mx) / (2 + mx)
    qx[100] = 1.0
    return qx


def smoothed_estimated_high_ages(
    estimated_2023_qx: dict[int, float],
    observed_qx: dict[tuple[str, int, int], float],
    ages: range,
) -> dict[int, float]:
    result: dict[int, float] = {}
    for age in ages:
        v2017 = observed_qx[("male", age, 2017)]
        v2018 = observed_qx[("male", age, 2018)]
        v2019 = observed_qx[("male", age, 2019)]
        v2022 = observed_qx[("male", age, 2022)]
        v2023 = estimated_2023_qx[age]

        prime_2023 = (v2019 + v2022 + 3 * v2023) / 5
        prime_2022 = (v2018 + v2019 + v2022 + 2 * v2023) / 5
        prime_2019 = (v2017 + v2018 + v2019 + v2022 + v2023) / 5
        result[age] = (prime_2019 + prime_2022 + 3 * prime_2023) / 5
    return result


def fit_alpha_beta(years: np.ndarray, e0_values: np.ndarray, e0_min: float, e0_max: float) -> tuple[float, float]:
    logit_values = np.log((e0_max - e0_values) / (e0_values - e0_min))
    design = np.column_stack([np.ones_like(years, dtype=float), years.astype(float)])
    alpha, beta = np.linalg.lstsq(design, logit_values, rcond=None)[0]
    return float(alpha), float(beta)


def project_e0(years: np.ndarray, alpha: float, beta: float, e0_min: float, e0_max: float) -> np.ndarray:
    logit_hat = alpha + beta * years
    return e0_min + (e0_max - e0_min) / (1 + np.exp(logit_hat))


def solve_e0_max_for_target(
    years: np.ndarray,
    e0_values: np.ndarray,
    e0_min: float,
    target_year: int,
    target_e0: float,
) -> float | None:
    lower = max(float(np.max(e0_values)) + 0.05, target_e0 + 0.05, e0_min + 0.1)
    upper = max(lower + 5.0, 120.0)

    def horizon_error(e0_max: float) -> float:
        if e0_max <= lower:
            return -1.0
        alpha, beta = fit_alpha_beta(years, e0_values, e0_min, e0_max)
        projected = project_e0(np.array([target_year], dtype=float), alpha, beta, e0_min, e0_max)[0]
        return float(projected - target_e0)

    error_low = horizon_error(lower)
    error_high = horizon_error(upper)
    attempts = 0
    while error_low * error_high > 0 and attempts < 25:
        upper += 20.0
        error_high = horizon_error(upper)
        attempts += 1

    if error_low * error_high > 0:
        return None

    lo = lower
    hi = upper
    for _ in range(80):
        mid = (lo + hi) / 2
        error_mid = horizon_error(mid)
        if abs(error_mid) < 1e-10:
            return mid
        if error_low * error_mid <= 0:
            hi = mid
            error_high = error_mid
        else:
            lo = mid
            error_low = error_mid
    return (lo + hi) / 2


def estimate_asymptotes(
    sex: str,
    observed_e0: dict[tuple[str, int], float],
    target_year: int,
    target_e0: float,
) -> tuple[float, float, float, float]:
    years = np.array(
        [
            year
            for year in sorted(year for current_sex, year in observed_e0 if current_sex == sex)
            if OBSERVED_E0_START_YEAR <= year <= LAST_OBSERVED_YEAR and year not in EXCLUDED_REGRESSION_YEARS
        ],
        dtype=float,
    )
    e0_values = np.array([observed_e0[(sex, int(year))] for year in years], dtype=float)

    best: tuple[float, float, float, float] | None = None
    for e0_min in np.arange(40.0, 70.01, 0.05):
        if e0_min >= float(np.min(e0_values)) - 0.05:
            continue
        e0_max = solve_e0_max_for_target(years, e0_values, e0_min, target_year, target_e0)
        if e0_max is None:
            continue
        alpha, beta = fit_alpha_beta(years, e0_values, e0_min, e0_max)
        fitted = project_e0(years, alpha, beta, e0_min, e0_max)
        sse = float(np.sum((fitted - e0_values) ** 2))
        candidate = (sse, float(e0_min), float(e0_max), alpha, beta)
        if best is None or candidate < best:
            best = candidate

    if best is None:
        raise ValueError(f"Could not estimate asymptotes for {sex}")

    _, e0_min_best, e0_max_best, alpha_best, beta_best = best
    return e0_min_best, e0_max_best, alpha_best, beta_best


def fit_logit_projection(
    sex: str,
    observed_e0: dict[tuple[str, int], float],
    horizon_target_e0: float,
    gap_distribution_years: int,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    e0_min, e0_max, alpha, beta = estimate_asymptotes(
        sex,
        observed_e0,
        END_YEAR,
        horizon_target_e0,
    )

    projection_years = np.arange(OBSERVED_E0_START_YEAR, END_YEAR + 1, dtype=int)
    e0_hat = project_e0(projection_years.astype(float), alpha, beta, e0_min, e0_max)
    logit_hat = alpha + beta * projection_years

    e0_observed_last = observed_e0[(sex, LAST_OBSERVED_YEAR)]
    e0_hat_last = float(e0_hat[np.where(projection_years == LAST_OBSERVED_YEAR)][0])
    e0_gap_last = e0_observed_last - e0_hat_last

    gap_factor: list[float] = []
    for year in projection_years:
        if year == LAST_OBSERVED_YEAR:
            gap_factor.append(1.0)
        elif LAST_OBSERVED_YEAR < year <= (LAST_OBSERVED_YEAR - 1 + gap_distribution_years):
            gap_factor.append(1 - ((year - (LAST_OBSERVED_YEAR - 1)) / gap_distribution_years))
        else:
            gap_factor.append(0.0)

    gap_factor_np = np.array(gap_factor, dtype=float)
    e0_adjusted = e0_hat + e0_gap_last * gap_factor_np

    coef_t: list[float | None] = []
    for idx, year in enumerate(projection_years):
        if year < START_YEAR:
            coef_t.append(None)
        elif year == END_YEAR:
            coef_t.append(1.0)
        else:
            prev = e0_adjusted[idx - 1]
            coef_t.append((e0_adjusted[idx] - prev) / (horizon_target_e0 - prev))

    e0_projection = pd.DataFrame(
        {
            "Year": projection_years,
            "e0_observed": [observed_e0.get((sex, int(year))) for year in projection_years],
            "logit_e0_hat": logit_hat,
            "e0_hat": e0_hat,
            "gap_factor": gap_factor_np,
            "e0_adjusted": e0_adjusted,
            "coef_t": coef_t,
            "alpha": alpha,
            "beta": beta,
            "e0_max": e0_max,
            "e0_min": e0_min,
            "e0_gap_last_observed": e0_gap_last,
            "e0_horizon_target": horizon_target_e0,
        }
    )

    parameters = pd.DataFrame(
        [
            {"parameter": "sex", "value": sex},
            {"parameter": "regression_start_year", "value": OBSERVED_E0_START_YEAR},
            {"parameter": "regression_end_year", "value": LAST_OBSERVED_YEAR},
            {"parameter": "excluded_regression_years", "value": ",".join(str(year) for year in sorted(EXCLUDED_REGRESSION_YEARS))},
            {"parameter": "gap_distribution_years", "value": gap_distribution_years},
            {"parameter": "e0_horizon_target", "value": horizon_target_e0},
            {"parameter": "e0_min", "value": e0_min},
            {"parameter": "e0_max", "value": e0_max},
            {"parameter": "alpha", "value": alpha},
            {"parameter": "beta", "value": beta},
        ]
    )

    return e0_projection, parameters


def parse_model_qx_profiles(sex: str, family: str, lower_e0: int, upper_e0: int) -> dict[int, dict[int, float]]:
    wb = load_workbook(MODEL_TABLE_FILE, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: pos for pos, name in enumerate(header)}

    profiles: dict[int, dict[int, float]] = {lower_e0: {}, upper_e0: {}}
    sex_label = sex.title()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx["Type_MLT"]] != "CD":
            continue
        if row[idx["Family"]] != family:
            continue
        if row[idx["Sex"]] != sex_label:
            continue
        e0 = int(row[idx["E0"]])
        if e0 not in profiles:
            continue
        age = int(row[idx["age"]])
        profiles[e0][age] = float(row[idx["qx1"]])

    return profiles


def parse_model_ax_profiles(
    sex: str,
    family: str,
    lower_e0: int,
    upper_e0: int,
) -> dict[int, dict[int, float]]:
    wb = load_workbook(MODEL_TABLE_FILE, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: pos for pos, name in enumerate(header)}

    rows: dict[int, dict[int, dict[str, float]]] = {lower_e0: {}, upper_e0: {}}
    sex_label = sex.title()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx["Type_MLT"]] != "CD":
            continue
        if row[idx["Family"]] != family:
            continue
        if row[idx["Sex"]] != sex_label:
            continue
        e0 = int(row[idx["E0"]])
        if e0 not in rows:
            continue

        age = int(row[idx["age"]])
        rows[e0][age] = {
            "lx": float(row[idx["lx1"]]),
            "dx": float(row[idx["dx1"]]),
            "Lx": float(row[idx["Lx1"]]),
            "Tx": float(row[idx["Tx1"]]),
        }

    profiles: dict[int, dict[int, float]] = {}
    for e0, by_age in rows.items():
        profile: dict[int, float] = {}
        for age in range(0, 100):
            dx = by_age[age]["dx"]
            if dx == 0:
                profile[age] = 0.5
            else:
                profile[age] = (by_age[age]["Lx"] - by_age[age + 1]["lx"]) / dx
        profile[100] = by_age[100]["Tx"] / by_age[100]["lx"]
        profiles[e0] = profile

    return profiles


def interpolate_profiles(
    lower_profile: dict[int, float],
    upper_profile: dict[int, float],
    coefficient: float,
) -> dict[int, float]:
    return {
        age: coefficient * upper_profile[age] + (1 - coefficient) * lower_profile[age]
        for age in sorted(lower_profile)
    }


def build_yearly_profiles(
    smoothed_2019: dict[int, float],
    horizon_profile: dict[int, float],
    coef_series: dict[int, float | None],
    *,
    force_terminal_qx: bool = False,
) -> pd.DataFrame:
    ages = sorted(smoothed_2019)
    previous = smoothed_2019.copy()
    projected: dict[int, dict[int, float]] = {}

    for year in range(START_YEAR, END_YEAR):
        coefficient = coef_series[year]
        if coefficient is None:
            raise ValueError(f"Missing coefficient for year {year}")
        current = {
            age: coefficient * horizon_profile[age] + (1 - coefficient) * previous[age]
            for age in ages
        }
        if force_terminal_qx and 100 in current:
            current[100] = 1.0
        projected[year] = current
        previous = current

    projected[END_YEAR] = {
        age: (1.0 if force_terminal_qx and age == 100 else value)
        for age, value in horizon_profile.items()
    }

    df = pd.DataFrame({"Age": ages})
    for year in range(START_YEAR, END_YEAR + 1):
        df[f"Year_{year}"] = [projected[year][age] for age in ages]
    return df


def life_expectancy_at_birth(qx_by_age: dict[int, float], ax_by_age: dict[int, float]) -> float:
    lx = 100000.0
    total_Lx = 0.0
    for age in range(0, 101):
        qx = 1.0 if age == 100 else max(0.0, min(1.0, qx_by_age[age]))
        dx = lx * qx
        next_lx = 0.0 if age == 100 else lx - dx
        total_Lx += next_lx + ax_by_age[age] * dx
        lx = next_lx
    return total_Lx / 100000.0


def parse_benchmark_qx() -> dict[str, dict[tuple[int, int], float]]:
    wb = load_workbook(BENCHMARK_FILE, read_only=True, data_only=True)
    ws = wb["tabla-36774"]

    header_years = [int(ws.cell(7, col).value) for col in range(2, ws.max_column + 1)]
    benchmark: dict[str, dict[tuple[int, int], float]] = {"male": {}, "female": {}}

    current_sex: str | None = None
    for row in ws.iter_rows(min_row=8, values_only=True):
        label = row[0]
        if label is None:
            continue
        text = str(label).strip()
        if text == "Males":
            current_sex = "male"
            continue
        if text == "Females":
            current_sex = "female"
            continue

        if current_sex is None:
            continue

        match = re.search(r"\d+", text)
        if match is None:
            continue
        age = int(match.group())
        for idx, year in enumerate(header_years, start=1):
            value = row[idx]
            if value is None:
                continue
            benchmark[current_sex][(age, year)] = float(value)

    return benchmark


def parse_benchmark_e0() -> dict[str, dict[int, float]]:
    wb = load_workbook(BENCHMARK_E0_FILE, read_only=True, data_only=True)
    ws = wb["tabla-36775"]

    header_years = [int(ws.cell(7, col).value) for col in range(2, ws.max_column + 1)]
    benchmark: dict[str, dict[int, float]] = {"male": {}, "female": {}}

    current_sex: str | None = None
    for row in ws.iter_rows(min_row=8, values_only=True):
        label = row[0]
        if label is None:
            continue
        text = str(label).strip()
        if text == "Males":
            current_sex = "male"
            continue
        if text == "Females":
            current_sex = "female"
            continue

        if current_sex is None:
            continue

        if text != "0 years old":
            continue

        for idx, year in enumerate(header_years, start=1):
            value = row[idx]
            if value is None:
                continue
            benchmark[current_sex][year] = float(value)

    return benchmark


def compare_dicts(lhs: dict, rhs: dict) -> tuple[list[tuple], SummaryRow]:
    common_keys = sorted(set(lhs).intersection(rhs))
    rows: list[tuple] = []
    abs_errors: list[float] = []

    for key in common_keys:
        lhs_value = lhs[key]
        rhs_value = rhs[key]
        abs_error = abs(lhs_value - rhs_value)
        key_values = key if isinstance(key, tuple) else (key,)
        rows.append((*key_values, lhs_value, rhs_value, abs_error))
        abs_errors.append(abs_error)

    mean_abs_error = sum(abs_errors) / len(abs_errors) if abs_errors else math.nan
    max_abs_error = max(abs_errors) if abs_errors else math.nan
    summary = SummaryRow(
        sex="",
        metric="",
        compared_cells=len(abs_errors),
        mean_abs_error=mean_abs_error,
        max_abs_error=max_abs_error,
    )
    return rows, summary


def write_sheet(ws, headers: list[str], rows: list[tuple] | list[list] | list[dict]) -> None:
    ws.append(headers)
    for row in rows:
        if isinstance(row, dict):
            ws.append([row.get(header) for header in headers])
        else:
            ws.append(list(row))


def save_dataframe(df: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(path, index=False)


def run_for_sex(
    config: SexConfig,
    observed_qx: dict[tuple[str, int, int], float],
    observed_ax: dict[tuple[str, int, int], float],
    observed_e0: dict[tuple[str, int], float],
) -> dict[str, pd.DataFrame | dict]:
    e0_projection, parameters = fit_logit_projection(
        config.sex,
        observed_e0,
        config.horizon_target_e0,
        config.gap_distribution_years,
    )

    observed_qx_wide = build_observed_wide(observed_qx, config.sex, SMOOTHING_YEARS)
    observed_ax_wide = build_observed_wide(observed_ax, config.sex, SMOOTHING_YEARS)
    smoothed_qx = two_pass_moving_average(observed_qx_wide)
    smoothed_ax = two_pass_moving_average(observed_ax_wide)
    start_qx = build_current_start_profile(observed_qx, config.sex)
    start_ax = build_current_start_profile(observed_ax, config.sex)

    e0_horizon = float(e0_projection.loc[e0_projection["Year"] == END_YEAR, "e0_horizon_target"].iloc[0])
    lower_e0 = math.floor(e0_horizon)
    upper_e0 = lower_e0 + 1
    coef_horizon = (e0_horizon - lower_e0) / (upper_e0 - lower_e0)

    qx_profiles = parse_model_qx_profiles(config.sex, config.model_family, lower_e0, upper_e0)
    ax_profiles = parse_model_ax_profiles(config.sex, config.model_family, lower_e0, upper_e0)

    qx_horizon = interpolate_profiles(qx_profiles[lower_e0], qx_profiles[upper_e0], coef_horizon)
    qx_horizon[100] = 1.0
    ax_horizon = interpolate_profiles(ax_profiles[lower_e0], ax_profiles[upper_e0], coef_horizon)

    coef_series = {int(row["Year"]): row["coef_t"] for _, row in e0_projection.iterrows() if START_YEAR <= row["Year"] <= END_YEAR}

    qx_projected = build_yearly_profiles(
        {int(row.Age): float(row.smoothed_start) for row in start_qx.itertuples(index=False)},
        qx_horizon,
        coef_series,
        force_terminal_qx=True,
    )
    ax_projected = build_yearly_profiles(
        {int(row.Age): float(row.smoothed_start) for row in start_ax.itertuples(index=False)},
        ax_horizon,
        coef_series,
        force_terminal_qx=False,
    )

    life_rows: list[dict[str, float | int]] = []
    for year in range(START_YEAR, END_YEAR + 1):
        qx_by_age = {int(row.Age): float(getattr(row, f"Year_{year}")) for row in qx_projected.itertuples(index=False)}
        ax_by_age = {int(row.Age): float(getattr(row, f"Year_{year}")) for row in ax_projected.itertuples(index=False)}
        e0_from_life_table = life_expectancy_at_birth(qx_by_age, ax_by_age)
        life_rows.append({"Year": year, "e0_life_table": e0_from_life_table})
    life_projection = pd.DataFrame(life_rows)

    qx_horizon_df = pd.DataFrame({"Age": sorted(qx_horizon), "qx_proj2073": [qx_horizon[age] for age in sorted(qx_horizon)]})
    ax_horizon_df = pd.DataFrame({"Age": sorted(ax_horizon), "ax_proj2073": [ax_horizon[age] for age in sorted(ax_horizon)]})

    return {
        "parameters": parameters,
        "e0_projection": e0_projection,
        "observed_qx_wide": observed_qx_wide,
        "observed_ax_wide": observed_ax_wide,
        "smoothed_qx": smoothed_qx,
        "smoothed_ax": smoothed_ax,
        "start_qx": start_qx,
        "start_ax": start_ax,
        "qx_horizon": qx_horizon_df,
        "ax_horizon": ax_horizon_df,
        "qx_projected": qx_projected,
        "ax_projected": ax_projected,
        "life_projection": life_projection,
    }


def rebuild_result_with_start_profiles(
    result: dict[str, pd.DataFrame | dict],
    start_qx: pd.DataFrame,
    start_ax: pd.DataFrame,
) -> dict[str, pd.DataFrame | dict]:
    coef_series = {
        int(row["Year"]): row["coef_t"]
        for _, row in result["e0_projection"].iterrows()
        if START_YEAR <= row["Year"] <= END_YEAR
    }
    qx_horizon = {
        int(row.Age): float(row.qx_proj2073)
        for row in result["qx_horizon"].itertuples(index=False)
    }
    ax_horizon = {
        int(row.Age): float(row.ax_proj2073)
        for row in result["ax_horizon"].itertuples(index=False)
    }

    qx_projected = build_yearly_profiles(
        {int(row.Age): float(row.smoothed_start) for row in start_qx.itertuples(index=False)},
        qx_horizon,
        coef_series,
        force_terminal_qx=True,
    )
    ax_projected = build_yearly_profiles(
        {int(row.Age): float(row.smoothed_start) for row in start_ax.itertuples(index=False)},
        ax_horizon,
        coef_series,
        force_terminal_qx=False,
    )

    life_rows: list[dict[str, float | int]] = []
    for year in range(START_YEAR, END_YEAR + 1):
        qx_by_age = {int(row.Age): float(getattr(row, f"Year_{year}")) for row in qx_projected.itertuples(index=False)}
        ax_by_age = {int(row.Age): float(getattr(row, f"Year_{year}")) for row in ax_projected.itertuples(index=False)}
        e0_from_life_table = life_expectancy_at_birth(qx_by_age, ax_by_age)
        life_rows.append({"Year": year, "e0_life_table": e0_from_life_table})
    life_projection = pd.DataFrame(life_rows)

    rebuilt = dict(result)
    rebuilt["start_qx"] = start_qx
    rebuilt["start_ax"] = start_ax
    rebuilt["qx_projected"] = qx_projected
    rebuilt["ax_projected"] = ax_projected
    rebuilt["life_projection"] = life_projection
    return rebuilt


def build_all_results(
    observed_qx: dict[tuple[str, int, int], float],
    observed_ax: dict[tuple[str, int, int], float],
    observed_e0: dict[tuple[str, int], float],
) -> dict[str, dict[str, dict[str, pd.DataFrame | dict]]]:
    baseline_results = {
        config.sex: run_for_sex(config, observed_qx, observed_ax, observed_e0)
        for config in CONFIGS
    }
    results: dict[str, dict[str, dict[str, pd.DataFrame | dict]]] = {
        VARIANT_BASELINE: baseline_results
    }

    if POPULATION_EXTRACT_CSV.exists() and DEATHS_EXTRACT_CSV.exists():
        population = load_population_male_july_2023()
        deaths = load_deaths_male_2023()
        estimated_2023 = estimated_qx_2023_male(population, deaths, observed_qx)
        adjusted_high_age_start = smoothed_estimated_high_ages(estimated_2023, observed_qx, range(95, 100))

        male_baseline = baseline_results["male"]
        adjusted_start_qx = male_baseline["start_qx"].copy()
        for age, value in adjusted_high_age_start.items():
            adjusted_start_qx.loc[adjusted_start_qx["Age"] == age, "smoothed_start"] = value

        adjusted_male = rebuild_result_with_start_profiles(
            male_baseline,
            adjusted_start_qx,
            male_baseline["start_ax"].copy(),
        )

        results[VARIANT_MALE_HIGH_AGE_ADJUSTED] = {
            "male": adjusted_male,
            "female": baseline_results["female"],
        }

    return results


def build_validation(
    results: dict[str, dict[str, dict[str, pd.DataFrame | dict]]],
    benchmark_qx: dict[str, dict[tuple[int, int], float]],
    benchmark_e0: dict[str, dict[int, float]],
    observed_qx: dict[tuple[str, int, int], float],
) -> None:
    output_wb = Workbook()
    summary_ws = output_wb.active
    summary_ws.title = "summary"
    summary_ws.append(["variant", "sex", "metric", "compared_cells", "mean_abs_error", "max_abs_error"])
    summary_lookup: dict[tuple[str, str, str], SummaryRow] = {}

    for variant_name, variant_results in results.items():
        variant_prefix = VARIANT_SHEET_PREFIX.get(variant_name, variant_name[:4])
        for config in CONFIGS:
            result = variant_results[config.sex]
            qx_df = result["qx_projected"]
            life_df = result["life_projection"]
            e0_projection = result["e0_projection"]

            generated_qx = {}
            for row in qx_df.itertuples(index=False):
                age = int(row.Age)
                for year in range(START_YEAR, END_YEAR + 1):
                    generated_qx[(age, year)] = float(getattr(row, f"Year_{year}")) * 1000.0
            generated_e0_adjusted = {
                int(row.Year): float(row.e0_adjusted)
                for row in e0_projection.itertuples(index=False)
                if START_YEAR <= row.Year <= END_YEAR
            }
            generated_e0_life = {int(row.Year): float(row.e0_life_table) for row in life_df.itertuples(index=False)}

            qx_rows, qx_summary = compare_dicts(generated_qx, benchmark_qx[config.sex])
            qx_summary.sex = config.sex
            qx_summary.metric = "qx_official_per_thousand_all_years"
            summary_lookup[(variant_name, config.sex, qx_summary.metric)] = qx_summary
            summary_ws.append([variant_name, qx_summary.sex, qx_summary.metric, qx_summary.compared_cells, qx_summary.mean_abs_error, qx_summary.max_abs_error])

            for specific_year in [START_YEAR, 2030, 2050, END_YEAR]:
                _, year_summary = compare_dicts(
                    {key: value for key, value in generated_qx.items() if key[1] == specific_year},
                    {key: value for key, value in benchmark_qx[config.sex].items() if key[1] == specific_year},
                )
                year_summary.sex = config.sex
                year_summary.metric = f"qx_official_per_thousand_{specific_year}"
                summary_lookup[(variant_name, config.sex, year_summary.metric)] = year_summary
                summary_ws.append(
                    [
                        variant_name,
                        year_summary.sex,
                        year_summary.metric,
                        year_summary.compared_cells,
                        year_summary.mean_abs_error,
                        year_summary.max_abs_error,
                    ]
                )

            qx_ws = output_wb.create_sheet(f"{variant_prefix}_{config.sex}_qx")
            write_sheet(qx_ws, ["age", "year", "generated", "benchmark", "abs_error"], qx_rows)

            top_qx_ws = output_wb.create_sheet(f"{variant_prefix}_{config.sex}_top")
            top_qx_rows = sorted(qx_rows, key=lambda row: row[-1], reverse=True)[:200]
            write_sheet(top_qx_ws, ["age", "year", "generated", "benchmark", "abs_error"], top_qx_rows)

            coef_first_year = float(e0_projection.loc[e0_projection["Year"] == START_YEAR, "coef_t"].iloc[0])
            qx_horizon_by_age = {
                int(row.Age): float(row.qx_proj2073) * 1000.0
                for row in result["qx_horizon"].itertuples(index=False)
            }
            start_qx_by_age = {
                int(row.Age): float(row.smoothed_start) * 1000.0
                for row in result["start_qx"].itertuples(index=False)
            }
            start_diag_rows = []
            for age in range(0, 101):
                benchmark_first_year = benchmark_qx[config.sex][(age, START_YEAR)]
                implied_start = (benchmark_first_year - coef_first_year * qx_horizon_by_age[age]) / (1 - coef_first_year)
                start_diag_rows.append(
                    [
                        age,
                        observed_qx[(config.sex, age, LAST_OBSERVED_YEAR)] * 1000.0,
                        start_qx_by_age[age],
                        implied_start,
                        qx_horizon_by_age[age],
                        float(qx_df.loc[qx_df["Age"] == age, f"Year_{START_YEAR}"].iloc[0]) * 1000.0,
                        benchmark_first_year,
                        implied_start - start_qx_by_age[age],
                        implied_start - observed_qx[(config.sex, age, LAST_OBSERVED_YEAR)] * 1000.0,
                    ]
                )
            start_diag_ws = output_wb.create_sheet(f"{variant_prefix}_{config.sex}_diag")
            write_sheet(
                start_diag_ws,
                [
                    "age",
                    "observed_2023_qx_per_thousand",
                    "current_start_qx_per_thousand",
                    "implied_start_from_benchmark_2024",
                    "horizon_2073_qx_per_thousand",
                    "generated_2024_qx_per_thousand",
                    "benchmark_2024_qx_per_thousand",
                    "implied_minus_current_start",
                    "implied_minus_observed_2023",
                ],
                start_diag_rows,
            )

            life_ws = output_wb.create_sheet(f"{variant_prefix}_{config.sex}_e0")
            helper_rows = []
            for year in range(START_YEAR, END_YEAR + 1):
                helper_rows.append(
                    [
                        year,
                        generated_e0_adjusted[year],
                        generated_e0_life[year],
                        benchmark_e0[config.sex].get(year),
                        config.horizon_target_e0 if year == END_YEAR else None,
                    ]
                )
            write_sheet(
                life_ws,
                ["year", "generated_logit_e0", "generated_qx_projected_ax_e0", "official_e0_age0", "official_horizon_target_if_2073"],
                helper_rows,
            )

            _, e0_life_summary = compare_dicts(generated_e0_life, benchmark_e0[config.sex])
            e0_life_summary.sex = config.sex
            e0_life_summary.metric = "e0_life_table_vs_official_age0"
            summary_lookup[(variant_name, config.sex, e0_life_summary.metric)] = e0_life_summary
            summary_ws.append(
                [
                    variant_name,
                    e0_life_summary.sex,
                    e0_life_summary.metric,
                    e0_life_summary.compared_cells,
                    e0_life_summary.mean_abs_error,
                    e0_life_summary.max_abs_error,
                ]
            )

            _, e0_path_summary = compare_dicts(generated_e0_adjusted, benchmark_e0[config.sex])
            e0_path_summary.sex = config.sex
            e0_path_summary.metric = "e0_logit_path_vs_official_age0"
            summary_lookup[(variant_name, config.sex, e0_path_summary.metric)] = e0_path_summary
            summary_ws.append(
                [
                    variant_name,
                    e0_path_summary.sex,
                    e0_path_summary.metric,
                    e0_path_summary.compared_cells,
                    e0_path_summary.mean_abs_error,
                    e0_path_summary.max_abs_error,
                ]
            )

            e0_target_summary = SummaryRow(
                sex=config.sex,
                metric="e0_life_table_2073_vs_target",
                compared_cells=1,
                mean_abs_error=abs(generated_e0_life[END_YEAR] - config.horizon_target_e0),
                max_abs_error=abs(generated_e0_life[END_YEAR] - config.horizon_target_e0),
            )
            summary_lookup[(variant_name, config.sex, e0_target_summary.metric)] = e0_target_summary
            summary_ws.append(
                [
                    variant_name,
                    e0_target_summary.sex,
                    e0_target_summary.metric,
                    e0_target_summary.compared_cells,
                    e0_target_summary.mean_abs_error,
                    e0_target_summary.max_abs_error,
                ]
            )

    if VARIANT_MALE_HIGH_AGE_ADJUSTED in results:
        comparison_ws = output_wb.create_sheet("male_variant_compare")
        comparison_rows = []
        for metric in [
            "qx_official_per_thousand_all_years",
            f"qx_official_per_thousand_{START_YEAR}",
            "qx_official_per_thousand_2030",
            "qx_official_per_thousand_2050",
            f"qx_official_per_thousand_{END_YEAR}",
            "e0_life_table_vs_official_age0",
        ]:
            baseline_summary = summary_lookup[(VARIANT_BASELINE, "male", metric)]
            adjusted_summary = summary_lookup[(VARIANT_MALE_HIGH_AGE_ADJUSTED, "male", metric)]
            comparison_rows.append(
                [
                    metric,
                    baseline_summary.mean_abs_error,
                    adjusted_summary.mean_abs_error,
                    baseline_summary.mean_abs_error - adjusted_summary.mean_abs_error,
                    baseline_summary.max_abs_error,
                    adjusted_summary.max_abs_error,
                ]
            )
        write_sheet(
            comparison_ws,
            [
                "metric",
                "baseline_mean_abs_error",
                "adjusted_mean_abs_error",
                "improvement_in_mean_abs_error",
                "baseline_max_abs_error",
                "adjusted_max_abs_error",
            ],
            comparison_rows,
        )

    notes_ws = output_wb.create_sheet("method_notes")
    write_sheet(
        notes_ws,
        ["topic", "note"],
        [
            ["observed source", "The active 2024-2073 run uses the fresh official INE mortality-table workbook 27153 as the observed source."],
            ["benchmark source", "Comparisons use INE table 36774, Projected Mortality Tables 2024-2073: Risk of death by age and sex (per thousand)."],
            ["life-expectancy benchmark source", "Projected life expectancy at age 0 is validated against INE table 36775, Projected Mortality Tables 2024-2073: Life expectancy by age and sex."],
            ["current methodology", "The 2024 methodology uses 2023 as the last observed year for the e0 regression and excludes 2020 and 2021 from that regression."],
            ["horizon targets", "The 2073 life-expectancy targets are 86.0 years for men and 90.0 years for women, as stated in the official INE 2024-2074 methodology and press release."],
            ["replication assumption", "For the age profile interpolation, this rough replication uses a pandemic-adjusted starting profile built from 2017, 2018, 2019, 2022 and 2023, which better approximates the current methodology than carrying forward the old 2019-only start profile."],
            ["optional variant", "The male_high_age_adjusted variant keeps the baseline methodology and only replaces male ages 95-99 in the starting qx profile with a proxy derived from 2023 deaths and 1 July 2023 population, then smooths those values before projection."],
        ],
    )

    output_wb.save(VALIDATION_FILE)


def save_outputs(results: dict[str, dict[str, dict[str, pd.DataFrame | dict]]]) -> None:
    for variant_name, variant_results in results.items():
        suffix = "" if variant_name == VARIANT_BASELINE else f"_{variant_name}"
        for config in CONFIGS:
            if config.sex not in variant_results:
                continue
            if variant_name != VARIANT_BASELINE and config.sex != "male":
                continue

            result = variant_results[config.sex]
            save_dataframe(result["parameters"], INTERMEDIATE_DIR / f"ine_parameters_{config.sex}_2024_2073{suffix}.xlsx")
            save_dataframe(result["e0_projection"], INTERMEDIATE_DIR / f"ine_e0_projection_{config.sex}_2024_2073{suffix}.xlsx")
            save_dataframe(result["observed_qx_wide"], INTERMEDIATE_DIR / f"ine_observed_qx_2015_2019_{config.sex}.xlsx")
            save_dataframe(result["observed_ax_wide"], INTERMEDIATE_DIR / f"ine_observed_ax_2015_2019_{config.sex}.xlsx")
            save_dataframe(result["smoothed_qx"], INTERMEDIATE_DIR / f"ine_qx_2019_twice_smoothed_{config.sex}.xlsx")
            save_dataframe(result["smoothed_ax"], INTERMEDIATE_DIR / f"ine_ax_2019_twice_smoothed_{config.sex}.xlsx")
            save_dataframe(result["start_qx"], INTERMEDIATE_DIR / f"ine_qx_start_profile_{config.sex}_2017_2019_2022_2023{suffix}.xlsx")
            save_dataframe(result["start_ax"], INTERMEDIATE_DIR / f"ine_ax_start_profile_{config.sex}_2017_2019_2022_2023{suffix}.xlsx")
            save_dataframe(result["qx_horizon"], INTERMEDIATE_DIR / f"ine_qx_projected_2073_{config.sex}{suffix}.xlsx")
            save_dataframe(result["ax_horizon"], INTERMEDIATE_DIR / f"ine_ax_projected_2073_{config.sex}{suffix}.xlsx")
            save_dataframe(result["qx_projected"], FINAL_DIR / f"ine_qx_2024_2073_{config.sex}{suffix}.xlsx")
            save_dataframe(result["ax_projected"], FINAL_DIR / f"ine_ax_2024_2073_{config.sex}{suffix}.xlsx")
            save_dataframe(result["life_projection"], FINAL_DIR / f"ine_e0_life_table_{config.sex}_2024_2073{suffix}.xlsx")


def main() -> None:
    require_file(OBSERVED_FILE)
    require_file(MODEL_TABLE_FILE)
    require_file(BENCHMARK_FILE)
    require_file(BENCHMARK_E0_FILE)
    require_file(POPULATION_EXTRACT_CSV)
    require_file(DEATHS_EXTRACT_CSV)
    ensure_dirs()

    observed_qx, observed_ax, observed_e0 = parse_observed_ine_table()
    benchmark_qx = parse_benchmark_qx()
    benchmark_e0 = parse_benchmark_e0()
    results = build_all_results(observed_qx, observed_ax, observed_e0)
    save_outputs(results)
    build_validation(results, benchmark_qx, benchmark_e0, observed_qx)
    print(f"Wrote outputs to: {FINAL_DIR}")
    print(f"Wrote validation workbook: {VALIDATION_FILE}")


if __name__ == "__main__":
    main()
